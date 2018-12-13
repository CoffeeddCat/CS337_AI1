import numpy as np
# import xlrd
# import xlwt
import openpyxl


class Outer:

    def __init__(self, config, loader, network_upside, network_downside):
        self.config = config
        self.loader = loader
        self.network_upside = network_upside
        self.network_downside = network_downside

    def initialize(self):
        self.file = openpyxl.Workbook()
        self.sheet = self.file.active
        self.sheet.cell(row=1, column=1, value="Name")
        self.sheet.cell(row=1, column=2, value="Tooth_id")
        self.sheet.cell(row=1, column=3, value="Translation_X")
        self.sheet.cell(row=1, column=4, value="Translation_Y")
        self.sheet.cell(row=1, column=5, value="Translation_Z")
        self.sheet.cell(row=1, column=6, value="Rotation_X")
        self.sheet.cell(row=1, column=7, value="Rotation_Y")
        self.sheet.cell(row=1, column=8, value="Rotation_Z")
        self.sheet.cell(row=1, column=9, value="Rotation_W")

    def out(self, file_name):
        self.initialize()
        name_array, input_upside, input_downside = self.loader.give_all()
        upside_result = self.network_upside.return_mat(input_upside)
        downside_result = self.network_downside.return_mat(input_downside)
        print(upside_result)
        print(downside_result)
        now = 1
        for name_index in range(len(name_array)):
            for i in range(1, 33):
                self.sheet.cell(row=now + i, column=1,
                                value=name_array[name_index])
                self.sheet.cell(row=now + i, column=2, value=str(i))
                for j in range(3, 10):
                    if i <= 16:
                        self.sheet.cell(
                            row=now + i, column=j, value=str(upside_result[name_index][(i - 1) * 7 + j - 3]))
                    else:
                        self.sheet.cell(
                            row=now + i, column=j, value=str(downside_result[name_index][(i - 17) * 7 + j - 3]))
            now = now + 32
        self.file.save(filename=file_name)